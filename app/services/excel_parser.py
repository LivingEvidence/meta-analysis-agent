import logging

import openpyxl

log = logging.getLogger(__name__)


def parse_outcomes(excel_path: str) -> dict:
    """Parse the Outcomes index sheet from an Excel file.

    Returns a dict with keys ``outcomes`` (list of dicts) and ``all_sheets``
    (list of sheet names).
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Find Outcomes sheet (case-insensitive)
    outcomes_sheet = None
    for name in wb.sheetnames:
        if name.lower() == "outcomes":
            outcomes_sheet = wb[name]
            break

    if outcomes_sheet is None:
        wb.close()
        return {"error": "No 'Outcomes' sheet found", "available_sheets": wb.sheetnames, "outcomes": []}

    rows = list(outcomes_sheet.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return {"error": "Outcomes sheet is empty", "outcomes": [], "all_sheets": wb.sheetnames}

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    outcomes = []
    for row in rows[1:]:
        if not any(row):
            continue
        entry = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                entry[headers[i]] = val
        outcomes.append(entry)

    result = {"outcomes": outcomes, "all_sheets": wb.sheetnames}
    wb.close()
    log.info("Parsed %d outcomes from %s", len(outcomes), excel_path)
    return result
